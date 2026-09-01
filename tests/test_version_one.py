from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl
import pandas as pd
import yaml

from data_profile_tool.cli import main


def write_csv(path: Path, rows: list[list[object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle).writerows(rows)


def one_output(root: Path, project: str, kind: str) -> Path:
    outputs = list((root / "Projects" / project / "outputs" / kind).iterdir())
    assert len(outputs) == 1
    return outputs[0]


def test_blind_then_configured_csv(tmp_path: Path) -> None:
    source = tmp_path / "weather.csv"
    write_csv(source, [["station", "temperature", "note"], ["A", 10, ""], ["A", 12, " ok "], ["B", 14, "   "]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "weather"]) == 0
    project = tmp_path / "Projects" / "weather"
    config_path = project / "dataset-config.yaml"
    config = yaml.safe_load(config_path.read_text())
    assert config["reviewed"] is False
    assert config["profiling"]["categorical_threshold"] == 50
    assert config["tables"]["weather"]["columns"]["station"]["in_active_source"] is True
    assert "present" not in config["tables"]["weather"]["columns"]["station"]
    assert config["tables"]["weather"]["columns"]["station"]["role"] == "categorical"
    assert (project / config["source"]["preserved_path"]).read_bytes() == source.read_bytes()
    blind = one_output(tmp_path, "weather", "blind")
    workbook_path = blind / "weather_blind_profile.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    assert workbook.sheetnames == ["Dataset Overview", "Column Summary", "Attribute Breakdown"]
    column_sheet = workbook["Column Summary"]
    assert column_sheet.freeze_panes == "D2"
    assert [cell.value for cell in next(column_sheet.iter_rows(max_row=1))[:6]] == [
        "source_table", "column_name", "analytical_role", "physical_type", "row_count", "unique_count"
    ]
    assert column_sheet.sheet_view.showGridLines is not False
    metadata = json.loads((blind / "run-metadata.json").read_text())
    assert re.search(r"[+-]\d{2}:\d{2}$", metadata["created_at_local"])
    assert metadata["files"][0] == "weather_blind_profile.xlsx"
    config["reviewed"] = True
    config["tables"]["weather"]["columns"]["temperature"]["role"] = "continuous"
    config_path.write_text(yaml.safe_dump(config, sort_keys=False), encoding="utf-8")
    assert main(["--projects-root", str(tmp_path), "configured", "weather"]) == 0
    configured = one_output(tmp_path, "weather", "configured")
    assert (configured / "weather_configured_profile.xlsx").exists()
    summary = pd.read_csv(configured / "column-summary.csv")
    temperature = summary.loc[summary.column_name == "temperature"].iloc[0]
    assert temperature["mean"] == 12
    assert temperature["minimum"] == 10
    assert temperature["maximum"] == 14


def test_continuous_categorical_role_creates_both_summaries(tmp_path: Path) -> None:
    source = tmp_path / "people.csv"
    write_csv(source, [["age"], [20], [20], [30], [40], [None]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "people"]) == 0
    config_path = tmp_path / "Projects" / "people" / "dataset-config.yaml"
    config = yaml.safe_load(config_path.read_text())
    config["reviewed"] = True
    config["tables"]["people"]["columns"]["age"]["role"] = "continuous_categorical"
    config_path.write_text(yaml.safe_dump(config, sort_keys=False), encoding="utf-8")

    assert main(["--projects-root", str(tmp_path), "configured", "people"]) == 0
    output = one_output(tmp_path, "people", "configured")
    columns = pd.read_csv(output / "column-summary.csv")
    age = columns.loc[columns.column_name == "age"].iloc[0]
    assert age["analytical_role"] == "continuous_categorical"
    assert age["mean"] == 27.5
    attributes = pd.read_csv(output / "attribute-breakdown.csv")
    assert set(attributes.loc[attributes.column_name == "age", "value"]) == {"20", "30", "40", "(empty string)"}


def test_configured_profile_warns_without_blocking_or_truncating(tmp_path: Path, capsys) -> None:
    source = tmp_path / "warning_data.csv"
    write_csv(source, [["numeric_mismatch", "large_category"], *[["not numeric", f"value-{value}"] for value in range(60)]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "warnings"]) == 0
    config_path = tmp_path / "Projects" / "warnings" / "dataset-config.yaml"
    config = yaml.safe_load(config_path.read_text())
    config["reviewed"] = True
    columns = config["tables"]["warning_data"]["columns"]
    columns["numeric_mismatch"]["role"] = "continuous"
    columns["large_category"]["role"] = "categorical"
    config_path.write_text(yaml.safe_dump(config, sort_keys=False), encoding="utf-8")

    assert main(["--projects-root", str(tmp_path), "configured", "warnings"]) == 0
    terminal = capsys.readouterr()
    assert "none of its 60 non-null values could be interpreted as numbers" in terminal.err
    assert "has 60 distinct non-null values" in terminal.err
    assert "output was still created" in terminal.err
    output = one_output(tmp_path, "warnings", "configured")
    attributes = pd.read_csv(output / "attribute-breakdown.csv")
    assert len(attributes.loc[attributes.column_name == "large_category"]) == 60


def test_existing_project_refuses_and_explains_options(tmp_path: Path, capsys) -> None:
    source = tmp_path / "data.csv"; write_csv(source, [["x"], [1]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "demo"]) == 0
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "demo"]) == 2
    error = capsys.readouterr().err
    assert "Nothing was overwritten" in error and "You may choose" in error and "reblind" in error


def test_reblind_preserves_history_and_flags_schema_change(tmp_path: Path) -> None:
    first = tmp_path / "first.csv"; second = tmp_path / "second.csv"
    write_csv(first, [["x"], [1]]); write_csv(second, [["x", "y"], [2, "new"]])
    assert main(["--projects-root", str(tmp_path), "blind", str(first), "--name", "demo"]) == 0
    config_path = tmp_path / "Projects" / "demo" / "dataset-config.yaml"
    config = yaml.safe_load(config_path.read_text()); config["reviewed"] = True
    config_path.write_text(yaml.safe_dump(config, sort_keys=False), encoding="utf-8")
    assert main(["--projects-root", str(tmp_path), "reblind", "demo", str(second)]) == 0
    input_versions = list((tmp_path / "Projects" / "demo" / "input").iterdir())
    assert len(input_versions) == 2
    assert all(re.fullmatch(r"\d{4}_\d{2}_\d{2}_\d{2}_\d{2}_\d{2}_[A-Za-z0-9+_-]+", version.name) for version in input_versions)
    assert yaml.safe_load(config_path.read_text())["reviewed"] is False


def test_reblind_refreshes_unreviewed_roles_when_threshold_changes(tmp_path: Path) -> None:
    source = tmp_path / "cars.csv"
    write_csv(source, [["city-mpg"], *[[value] for value in range(30)]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "cars"]) == 0
    config_path = tmp_path / "Projects" / "cars" / "dataset-config.yaml"
    initial = yaml.safe_load(config_path.read_text())
    assert initial["tables"]["cars"]["columns"]["city-mpg"]["role"] == "categorical"

    assert main(["--projects-root", str(tmp_path), "reblind", "cars", "--categorical-threshold", "24"]) == 0
    updated = yaml.safe_load(config_path.read_text())
    column = updated["tables"]["cars"]["columns"]["city-mpg"]
    assert column["inferred_role"] == "continuous"
    assert column["role"] == "continuous"


def test_reblind_preserves_roles_from_reviewed_config(tmp_path: Path) -> None:
    source = tmp_path / "data.csv"
    write_csv(source, [["code"], *[[value] for value in range(30)]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "demo"]) == 0
    config_path = tmp_path / "Projects" / "demo" / "dataset-config.yaml"
    config = yaml.safe_load(config_path.read_text())
    config["reviewed"] = True
    config["tables"]["data"]["columns"]["code"]["role"] = "identifier"
    config_path.write_text(yaml.safe_dump(config, sort_keys=False), encoding="utf-8")

    assert main(["--projects-root", str(tmp_path), "reblind", "demo", "--categorical-threshold", "24"]) == 0
    updated = yaml.safe_load(config_path.read_text())
    column = updated["tables"]["data"]["columns"]["code"]
    assert column["inferred_role"] == "continuous"
    assert column["role"] == "identifier"


def test_excel_hidden_and_failed_sheets_are_visible_in_overview(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    workbook = openpyxl.Workbook(); good = workbook.active; good.title = "Daily"
    good.append(["day", "value"]); good.append(["Mon", 1])
    hidden = workbook.create_sheet("Notes"); hidden.append(["note"]); hidden.sheet_state = "hidden"
    bad = workbook.create_sheet("Broken"); bad.append(["", "x"]); bad.append([1, 2]); workbook.save(source)
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "weather_book"]) == 0
    overview = pd.read_csv(one_output(tmp_path, "weather_book", "blind") / "dataset-overview.csv")
    assert dict(zip(overview.source_table, overview.status)) == {"Daily": "included", "Notes": "excluded", "Broken": "failed"}
    assert overview.loc[overview.source_table == "Notes", "exclusion_reason"].iloc[0] == "Hidden sheet"


def test_duplicate_csv_headers_fail_but_preserve_input(tmp_path: Path) -> None:
    source = tmp_path / "bad.csv"; write_csv(source, [["x", "x"], [1, 2]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "bad"]) == 2
    project = tmp_path / "Projects" / "bad"
    assert list((project / "input").glob("*/*"))
    assert (project / "ERROR.txt").exists()


def test_special_headers_do_not_break_or_become_formulas(tmp_path: Path) -> None:
    source = tmp_path / "special.csv"
    write_csv(source, [["pay ($/hr)", "=suspicious", "line\nbreak"], [20, "=1+1", "ok"]])
    assert main(["--projects-root", str(tmp_path), "blind", str(source), "--name", "special"]) == 0
    summary = pd.read_csv(one_output(tmp_path, "special", "blind") / "column-summary.csv")
    assert set(summary.column_name) == {"pay ($/hr)", "=suspicious", "line\nbreak"}
